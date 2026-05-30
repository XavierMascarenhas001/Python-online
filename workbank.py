import pandas as pd
import numpy as np

from tkinter import Tk, filedialog

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)

from openpyxl.drawing.image import Image

# =====================================================
# IMAGE PATHS
# =====================================================
IMG_LEFT = r"C:\Users\Xavier.Mascarenhas\OneDrive - Gaeltec Utilities Ltd\Desktop\Gaeltec\06_Programs\Dashboard\Images\GaeltecImage.png"

IMG_RIGHT = r"C:\Users\Xavier.Mascarenhas\OneDrive - Gaeltec Utilities Ltd\Desktop\Gaeltec\06_Programs\Dashboard\Images\SPEN.png"

# =====================================================
# FILE INPUTS
# =====================================================
root = Tk()
root.withdraw()

tracker_file = filedialog.askopenfilename(
    title="Select Project Tracker.parquet"
)

data_file = filedialog.askopenfilename(
    title="Select NEW BUILD PARQUET"
)

output_parquet = filedialog.asksaveasfilename(
    title="Save Updated Parquet",
    defaultextension=".parquet"
)

# =====================================================
# LOAD DATA
# =====================================================
tracker_df = pd.read_parquet(tracker_file)
df = pd.read_parquet(data_file)

# =====================================================
# CLEAN COLUMN NAMES
# =====================================================
tracker_df.columns = tracker_df.columns.str.strip().str.lower()
df.columns = df.columns.str.strip().str.lower()

# =====================================================
# TEXT CLEANER
# =====================================================
def clean_text(series):

    return (
        series.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'[^a-z0-9\s]', '', regex=True)
        .str.replace(r'\s+', ' ', regex=True)
    )

# =====================================================
# NORMALISE TEXT COLUMNS
# =====================================================
if 'job name' in tracker_df.columns:
    tracker_df['job_name_clean'] = clean_text(
        tracker_df['job name']
    )
else:
    tracker_df['job_name_clean'] = ""

if 'segment' in df.columns:
    df['segment_clean'] = clean_text(
        df['segment']
    )
else:
    df['segment_clean'] = ""

# =====================================================
# NORMALISE PID
# =====================================================
tracker_df['pid'] = tracker_df['pid'].astype(str).str.strip()
df['pid'] = df['pid'].astype(str).str.strip()

# =====================================================
# NORMALISE DATES
# =====================================================
df['datetouse'] = pd.to_datetime(
    df.get('datetouse'),
    errors='coerce'
)

df['plan1'] = pd.to_datetime(
    df.get('plan1'),
    errors='coerce'
)

# =====================================================
# ENSURE NUMERIC TOTAL
# =====================================================
df['total'] = pd.to_numeric(
    df.get('total'),
    errors='coerce'
).fillna(0)

# =====================================================
# POLE LOGIC
# =====================================================
def pole_value(x):

    x = str(x).lower()

    if "section structure 'h'" in x:
        return 2

    elif (
        "hv/ehv pole" in x
        or "lv structure single pole" in x
    ):
        return 1

    return 0

df['poles_calc'] = df['item'].apply(pole_value)

# =====================================================
# FILTER LAST YEAR
# =====================================================
last_year = df['datetouse'].dropna().dt.year.max()

df_last_year = df[
    df['datetouse'].dt.year == last_year
].copy()

# =====================================================
# QUARTERS
# =====================================================
df_last_year['quarter'] = df_last_year['datetouse'].dt.month.map({

    1: "Quarter 1",
    2: "Quarter 1",
    3: "Quarter 1",

    4: "Quarter 2",
    5: "Quarter 2",
    6: "Quarter 2",

    7: "Quarter 3",
    8: "Quarter 3",
    9: "Quarter 3",

    10: "Quarter 4",
    11: "Quarter 4",
    12: "Quarter 4",
})

# =====================================================
# PID SUMMARY
# =====================================================
group = df.groupby('pid')

summary = pd.DataFrame()

# TOTAL JOB VALUE
summary['total job value'] = group['total'].sum()

# UNPLANNED VALUE
summary['unplanned value'] = group.apply(
    lambda x: x.loc[
        x['datetouse'].isna(),
        'total'
    ].sum()
)

# TOTAL POLES
summary['total poles'] = group['poles_calc'].sum()

# POLES UNPLANNED
summary['poles unplanned'] = group.apply(
    lambda x: x.loc[
        x['datetouse'].isna(),
        'poles_calc'
    ].sum()
)

# LAST PLANNED WORK
summary['last planned work'] = group['plan1'].max()

# MOST COMMON COMMENT
summary['reason'] = group['comment'].agg(
    lambda x: (
        x.mode().iloc[0]
        if not x.mode().empty
        else np.nan
    )
)

# =====================================================
# QUARTER BREAKDOWN
# =====================================================
quarter_pivot = (
    df_last_year
    .groupby(['pid', 'quarter'])['poles_calc']
    .sum()
    .unstack(fill_value=0)
)

# ensure all quarters exist
for q in [
    "Quarter 1",
    "Quarter 2",
    "Quarter 3",
    "Quarter 4"
]:
    if q not in quarter_pivot.columns:
        quarter_pivot[q] = 0

quarter_pivot = quarter_pivot[
    [
        "Quarter 1",
        "Quarter 2",
        "Quarter 3",
        "Quarter 4"
    ]
]

# =====================================================
# MERGE SUMMARY + QUARTERS
# =====================================================
final = summary.join(
    quarter_pivot,
    how='left'
).fillna(0)

final = final.reset_index()

# =====================================================
# FIRST MERGE BY PID
# =====================================================
tracker_df = tracker_df.merge(
    final,
    on='pid',
    how='left',
    suffixes=('', '_new')
)

# =====================================================
# SECONDARY MATCH
# =====================================================
missing_mask = tracker_df['total job value'].isna()

segment_lookup = final.merge(
    df[['pid', 'segment_clean']].drop_duplicates(),
    on='pid',
    how='left'
)

segment_lookup = segment_lookup.drop_duplicates(
    subset=['segment_clean']
)

fallback_merge = tracker_df.loc[missing_mask].merge(
    segment_lookup,
    left_on='job_name_clean',
    right_on='segment_clean',
    how='left',
    suffixes=('', '_fallback')
)

# =====================================================
# COLUMNS TO FILL
# =====================================================
fill_cols = [

    'total job value',
    'unplanned value',
    'total poles',
    'poles unplanned',
    'last planned work',
    'reason',

    'Quarter 1',
    'Quarter 2',
    'Quarter 3',
    'Quarter 4'
]

# =====================================================
# FILL FROM FALLBACK
# =====================================================
for col in fill_cols:

    fallback_col = f"{col}_fallback"

    if fallback_col in fallback_merge.columns:

        tracker_df.loc[
            missing_mask,
            col
        ] = fallback_merge[fallback_col].values

# =====================================================
# DROP TEMP COLUMNS
# =====================================================
tracker_df = tracker_df.drop(
    columns=[
        'job_name_clean'
    ],
    errors='ignore'
)

tracker_df = tracker_df.drop(
    columns=[
        col for col in tracker_df.columns
        if col.endswith('_new')
    ],
    errors='ignore'
)

# =====================================================
# SAVE PARQUET
# =====================================================
tracker_df.to_parquet(
    output_parquet,
    index=False
)

# =====================================================
# EXCEL HELPERS
# =====================================================
def prepare_for_excel(df):

    df = df.copy()

    for col in df.columns:

        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.date

    return df

def format_excel_dates(df):

    return df

# =====================================================
# WRITE SHEET
# =====================================================
def write_sheet(wb, name, df):

    ws = wb.create_sheet(name[:31])

    if df is None or df.empty:
        ws.append(["No Data"])
        return

    df = prepare_for_excel(df)
    df = format_excel_dates(df)

    # FIX EXCEL <NA> ERRORS
    df = df.replace({pd.NA: None, np.nan: None})

    # =================================================
    # IMAGES
    # =================================================
    try:

        img1 = Image(IMG_LEFT)
        img2 = Image(IMG_RIGHT)

        img1.width = 110
        img1.height = 75

        img2.width = 140
        img2.height = 60

        ws.add_image(img1, "A1")
        ws.add_image(img2, "B1")

        ws.row_dimensions[1].height = 60

    except Exception as e:
        print(f"Image load failed: {e}")

    # =================================================
    # STYLES
    # =================================================
    HEADER_COLOR = "00CCFF"

    header_fill = PatternFill(
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        size=15
    )

    normal_font = Font(size=11)

    center_align = Alignment(
        horizontal="center"
    )

    currency_format = '£#,##0.00'

    thick_bottom = Border(
        bottom=Side(style="medium")
    )

    header_row = 2

    # =================================================
    # HEADERS
    # =================================================
    for col_idx, col_name in enumerate(df.columns, 1):

        cell = ws.cell(
            row=header_row,
            column=col_idx,
            value=col_name
        )

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # =================================================
    # DATA
    # =================================================
    for r_idx, row in enumerate(
        df.itertuples(index=False),
        header_row + 1
    ):

        for c_idx, value in enumerate(row, 1):

            # FIX REMAINING PANDAS NULLS
            if pd.isna(value):
                value = None

            col_name = df.columns[c_idx - 1]

            cell = ws.cell(
                row=r_idx,
                column=c_idx,
                value=value
            )

            cell.font = normal_font

            # =========================================
            # CURRENCY FORMAT
            # =========================================
            if col_name.lower() in [
                'total job value',
                'unplanned value'
            ]:
                cell.number_format = currency_format

            # =========================================
            # BORDER
            # =========================================
            cell.border = thick_bottom

    # =================================================
    # AUTO WIDTH
    # =================================================
    for column_cells in ws.columns:

        max_length = 0

        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            try:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

# =====================================================
# SAVE EXCEL
# =====================================================
wb = Workbook()

default_sheet = wb.active
wb.remove(default_sheet)

write_sheet(
    wb,
    "Summary",
    tracker_df
)

excel_output = output_parquet.replace(
    ".parquet",
    ".xlsx"
)

wb.save(excel_output)

print("DONE ✔")
print(f"Excel saved: {excel_output}")
print(f"Parquet saved: {output_parquet}")
